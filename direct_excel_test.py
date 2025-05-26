"""Direct test of Excel parsing functionality."""
import sys
import os
import logging
from pathlib import Path

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)
logger = logging.getLogger(__name__)

# Add the src directory to the Python path
project_root = Path(__file__).parent
src_dir = project_root / 'src'
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(src_dir))

class SimpleExcelParser:
    """A simple Excel parser for testing."""
    
    def __init__(self, bank, quarter):
        self.bank = bank
        self.quarter = quarter
    
    def parse(self, file_path):
        """Parse an Excel file and return structured data."""
        import pandas as pd
        from openpyxl import load_workbook
        
        logger.info("Loading Excel file: %s", file_path)
        
        # Read the Excel file
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        
        result = {
            'bank': self.bank,
            'quarter': self.quarter,
            'file_path': str(file_path),
            'document_type': 'test',
            'content': {
                'tables': {},
                'metadata': {}
            }
        }
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            logger.info("Processing sheet: %s", sheet_name)
            ws = wb[sheet_name]
            
            # Convert worksheet to DataFrame
            data = ws.values
            cols = next(data)
            df = pd.DataFrame(data, columns=cols)
            
            # Store sheet data
            result['content']['tables'][sheet_name] = {
                'shape': {
                    'rows': len(df),
                    'columns': len(df.columns)
                },
                'columns': [{'name': col} for col in df.columns],
                'data': df.head().to_dict('records')
            }
        
        return result

def main():
    """Run the test."""
    try:
        # Create a test Excel file
        test_file = Path("test_excel.xlsx")
        logger.info("Creating test Excel file: %s", test_file.absolute())
        
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"
        ws.append(["Name", "Value"])
        ws.append(["Test 1", 100])
        ws.append(["Test 2", 200])
        wb.save(test_file)
        
        if not test_file.exists():
            logger.error("Failed to create test file")
            return 1
            
        # Test the parser
        logger.info("Testing SimpleExcelParser...")
        parser = SimpleExcelParser(bank="test_bank", quarter="Q1_2025")
        result = parser.parse(test_file)
        
        # Print results
        print("\n=== Parser Result ===")
        print(f"Bank: {result.get('bank')}")
        print(f"Quarter: {result.get('quarter')}")
        print(f"Document Type: {result.get('document_type')}")
        
        if 'content' in result and 'tables' in result['content']:
            print("\nTables found:")
            for sheet_name, sheet_data in result['content']['tables'].items():
                print(f"\nSheet: {sheet_name}")
                print(f"Rows: {sheet_data.get('shape', {}).get('rows')}")
                print(f"Columns: {sheet_data.get('shape', {}).get('columns')}")
                
                if 'columns' in sheet_data:
                    print("Columns:", [col.get('name') for col in sheet_data['columns']])
                
                if 'data' in sheet_data and sheet_data['data']:
                    print("First row of data:", sheet_data['data'][0])
        
        return 0
        
    except Exception as e:
        logger.error("Test failed: %s", str(e), exc_info=True)
        return 1
    finally:
        # Clean up
        try:
            if 'test_file' in locals() and test_file.exists():
                os.remove(test_file)
                logger.info("Cleaned up test file")
        except Exception as e:
            logger.error("Error cleaning up: %s", str(e))

if __name__ == "__main__":
    sys.exit(main())
